from django.shortcuts import render, get_object_or_404, redirect

from .models import Componente
from .models import Aeronave
from .models import Vuelo
from .models import Orden

from .forms import Componente_formulario

from django.urls import reverse
from django.urls import reverse_lazy
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator

from django.views.generic.base import TemplateView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView
from django.views.generic.edit import UpdateView
from django.views.generic.edit import DeleteView

from .forms import *

from openpyxl import Workbook
from django.http import HttpResponse


import io
from django.http import FileResponse
from reportlab.pdfgen import canvas

from easy_pdf.views import PDFTemplateView

from easy_pdf.views import PDFTemplateResponseMixin
from django.conf import settings

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required


# Create your views here.

# class StaffRequiredMixin(object):
# 	# Mixin requerira que el usuario sea miembro del staff
# 	@method_decorator(staff_member_required)
# 	def dispatch(self, request, *args, **kwargs):
# 		return super(StaffRequiredMixin, self).dispatch(request, *args, **kwargs)

# -- Componentes -- #

@method_decorator(staff_member_required, name='dispatch')
class componenteListView(LoginRequiredMixin,ListView):
	model =	Componente
	template_name = 'componente/componente_list.html'

@method_decorator(staff_member_required, name='dispatch')
class componenteDetailView(LoginRequiredMixin,DetailView):
	model =	Componente
	template_name = 'componente/componente_detail.html'

@method_decorator(staff_member_required, name='dispatch')
class componenteCreateView(LoginRequiredMixin,CreateView):
	model = Componente
	form_class = ComponenteCreateForm
	template_name = 'componente/componente_form.html'
	success_url = reverse_lazy('componente_list')

@method_decorator(staff_member_required, name='dispatch')
class componenteUpdateView(LoginRequiredMixin,UpdateView):
    model = Componente
    form_class = ComponenteUpdateForm
    template_name = 'componente/componente_update_form.html'

    def get_success_url(self):
    	return reverse_lazy('componente_detail', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class componenteDeleteView(LoginRequiredMixin,DeleteView):
	model = Componente
	template_name = 'componente/componente_confirm_delete.html'
	success_url = reverse_lazy('componente_list')

@method_decorator(staff_member_required, name='dispatch')		
class componentePDFDetailView(LoginRequiredMixin, PDFTemplateResponseMixin, DetailView):
	model = Componente
	template_name = 'componente/pdf_componente_detail.html'
	
	def get_context_data(self, **kwargs):
		return super(componentePDFDetailView, self).get_context_data(
				pagesize='A4 landscape',
				title='Listado_aeronaves',
				encoding =u"utf-8",
				**kwargs
				)
@login_required
def componenteDownExcel(request):

	componentes = Componente.objects.all()

	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition']='attachment; filename="componentes.csv"'
	writer = csv.writer(response, delimiter=';')

	writer.writerow(['No. serie','Nombre', 'Marca'])

	for componente in componentes:
		writer.writerow([componente.numSerie, componente.nombre, componente.marca])

	return response


# **********************************************
@login_required
def componente_lista(request):
	componentes = Componente.objects.all()
	return render(request, 'componente/componente_lista.html', {'componentes': componentes})

@login_required
def componente_info(request, componente_id):
	componente = get_object_or_404(Componente, id=componente_id)
	return render(request, 'componente/componente_info.html', {'componente':componente})

@login_required
def componente_nuevo(request):
	componente_form = Componente_formulario()
	if request.method == "POST":
		componente_form = Componente_formulario(data=request.POST)
		if componente_form.is_valid():
			numSerie			= request.POST.get('numSerie', '')
			nombre				= request.POST.get('nombre', '')
			marca				= request.POST.get('marca', '')
			fabricante			= request.POST.get('fabricante', '')
			fechaFabricacion	= request.POST.get('fechaFabricacion', '')
			fechaVencimiento	= request.POST.get('fechaVencimiento', '')
			proveedor			= request.POST.get('proveedor', '')
			fechaIngreso		= request.POST.get('fechaingreso', '')
			descripcion			= request.POST.get('descripcion', '')
			return redirect(reverse('componente_nuevo')+"?componeteCreado")
	return render(request, 'componente/componente_nuevo.html', {'componente_form':componente_form})


# -- Aeronaves -- #

@method_decorator(staff_member_required, name='dispatch')
class aeronaveListView(LoginRequiredMixin,ListView):
	model =	Aeronave
	template_name = 'aeronave/aeronave_list.html'

@method_decorator(staff_member_required, name='dispatch')
class aeronaveDetailView(LoginRequiredMixin,DetailView):
	model =	Aeronave
	template_name = 'aeronave/aeronave_detail.html'

@method_decorator(staff_member_required, name='dispatch')
class aeronaveCreateView(LoginRequiredMixin,CreateView):
	model = Aeronave
	form_class = AeronaveCreateForm
	template_name = 'aeronave/aeronave_form.html'

	def form_valid(self, form):
		self.object = form.save(self)
		if self.object:
			
			ids = self.request.POST.getlist("componente")
			components = Componente.objects.filter(pk__in=ids)
			self.object.componente.clear()
			for idC in ids:
				for component in components:
					if str(component.pk) == idC:
						self.object.componente.add(component)
						break
			self.object.save()
			return super(aeronaveCreateView, self).form_valid(form)
		else:
			return self

	success_url = reverse_lazy('aeronave_list')

@method_decorator(staff_member_required, name='dispatch')
class aeronaveUpdateView(LoginRequiredMixin,UpdateView):
	model = Aeronave
	form_class = AeronaveUpdateForm
	template_name = 'aeronave/aeronave_update_form.html'

	def form_valid(self, form):
		self.object = form.save(self)
		if self.object:
			ids = self.request.POST.getlist("componente")
			components = Componente.objects.filter(pk__in=ids)
			self.object.componente.clear()
			for idC in ids:
				for component in components:
					if str(component.pk) == idC:
						self.object.componente.add(component)
						break
			self.object.save()
			return super(aeronaveUpdateView, self).form_valid(form)
		else:
			return self

	def get_success_url(self):
		return reverse_lazy('aeronave_detail', args=[self.object.id])+ '?ok'

@method_decorator(staff_member_required, name='dispatch')
class aeronaveDeleteView(LoginRequiredMixin,DeleteView):
	model = Aeronave
	template_name = 'aeronave/aeoronave_confirm_delete.html'
	success_url = reverse_lazy('aeronave_list')

@login_required
def aeronaveDownExcel(request):
	aeronaves = Aeronave.objects.all()

	response = HttpResponse(content_type='text/csv')
	response['Content-Disposition']='attachment; filename="aeronaves.csv"'
	writer = csv.writer(response, delimiter=';')
	
	writer.writerow(['Placa','Marca', 'Modelo', 'Componente(s)'])

	for aeronave in aeronaves:
		writer.writerow([aeronave.placa, aeronave.marca, aeronave.modelo, aeronave.componente])

	return response

@method_decorator(staff_member_required, name='dispatch')
class aeronaveReporteExcel(LoginRequiredMixin,TemplateView):
	def get(self, request, *args, **kwargs):
		aeronaves = Aeronave.objects.all()

		wb = Workbook()
		ws = wb.active
		ws['A1'] = 'REPORTE DE AERONAVES - DETALLES DE COMPONENTE'

		
		ws['A3'] = 'NO. SERIE'
		ws['A4'] = 'MARCA'
		ws['A5'] = 'PLACA'
		# ws['A7'] = 'COMPONENTES'
		# ws['A8'] = 'Nombre'
		# ws['B8'] = 'No. Serie'
		# ws['C8'] = 'Marca'
		# ws['D8'] = 'H. Util'
		# ws['E8'] = 'H. Utilizado'
		# ws['F8'] = 'H. Remanente'

		cont = 2
		for aeronave in aeronaves:
			ws.cell(row = 3, column = cont).value = aeronave.id
			ws.cell(row = 4, column = cont).value = aeronave.marca
			ws.cell(row = 5, column = cont).value = aeronave.placa
			cont+=1
			# for componente in aeronave.componente.all():
			# 	ws.cell(row = 9, column = 1).value = componente.nombre
			# 	ws.cell(row = 9, column = 2).value = componente.numSerie
			# 	ws.cell(row = 9, column = 3).value = componente.marca
			# 	ws.cell(row = 9, column = 4).value = componente.hvUtil
			# 	ws.cell(row = 9, column = 5).value = componente.hUtilizado
			# 	ws.cell(row = 9, column = 6).value = componente.horasRemanentes()



		nombre_archivo = 'aeronaveReporteExcel.xlsx'
		response = HttpResponse(content_type = "applications/ms-excel")
		content = "attachment; filename = {0}".format(nombre_archivo)
		response['Content-Disposition'] = content
		wb.save(response)

		return response

@method_decorator(staff_member_required, name='dispatch')
class aeronavePDFListView(LoginRequiredMixin,PDFTemplateView):
	template_name = 'aeronave/pdf_aeronave_list.html'

	def get_context_data(self, **kwargs):
		return super(aeronavePDFListView, self).get_context_data(
			pagesize='A4 landscape',
			title='Listado_aeronaves',
			**kwargs

			)
@method_decorator(staff_member_required, name='dispatch')		
class aeronavePDFDetailView(LoginRequiredMixin, PDFTemplateResponseMixin, DetailView):
	model = Aeronave
	template_name = 'aeronave/pdf_aeronave_detail.html'
	
	def get_context_data(self, **kwargs):
		return super(aeronavePDFDetailView, self).get_context_data(
			pagesize='A4 landscape',
			title='Detalle_aeronave',
			encoding =u"utf-8",
			**kwargs,
		)
# ***************************
@login_required
def aeronave_lista(request):
	aeronaves = Aeronave.objects.all()
	return render(request, 'aeronave/aeronave_lista.html', {'aeronaves':aeronaves})

@login_required
def aeronave_info(request, aeronave_id):
	aeronave = get_object_or_404(Aeronave, id=aeronave_id)
	return render(request, 'aeronave/aeronave_info.html',{'aeronave':aeronave})

# -- Vuelos -- #

@method_decorator(staff_member_required, name='dispatch')
class vueloListView(LoginRequiredMixin,ListView):
	model =	Vuelo
	template_name = 'vuelo/vuelo_list.html' 

@method_decorator(staff_member_required, name='dispatch')
class vueloDetailView(LoginRequiredMixin,DetailView):
	model =	Vuelo
	template_name = 'vuelo/vuelo_detail.html'

@method_decorator(staff_member_required, name='dispatch')
class vueloCreateView(LoginRequiredMixin,CreateView):
	model = Vuelo
	form_class = VueloCreateForm
	template_name = 'vuelo/vuelo_form.html'
	
	def form_valid(self, form):
		self.object = form.save(self)
		if self.object:
			horas = self.object.horas
			minutos = self.object.minutos
			for component in self.object.aeronave.componente.all():
				component.hUtilizado += horas
				component.mUtilizado += minutos
				component.hDurg += horas
				component.hRemanente = component.hvUtil - component.hDurg
				component.porcenUso = (component.hDurg * 100)/component.hvUtil
				if component.porcenUso >= 90:
					component.alerta = "Si"
				component.save()
			return super(vueloCreateView, self).form_valid(form)
		else:
			return self
	success_url = reverse_lazy('vuelo_list')

@method_decorator(staff_member_required, name='dispatch')
class vueloUpdateView(LoginRequiredMixin,UpdateView):
    model = Vuelo
    form_class = VueloUpdateForm
    template_name = 'vuelo/vuelo_update_form.html'
    def get_success_url(self):
    	return reverse_lazy('vuelo_detail', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class vueloDeleteView(LoginRequiredMixin,DeleteView):
	model = Vuelo
	template_name = 'vuelo/vuelo_confirm_delete.html'
	success_url = reverse_lazy('vuelo_list')

# -- Orden -- #

@method_decorator(staff_member_required, name='dispatch')
class ordenListView(LoginRequiredMixin,ListView):
	model =	Orden
	template_name = 'orden/orden_list.html'

@method_decorator(staff_member_required, name='dispatch')
class ordenDetailView(LoginRequiredMixin,DetailView):
	model =	Orden
	template_name = 'orden/orden_detail.html'

@method_decorator(staff_member_required, name='dispatch')
class ordenCreateView(LoginRequiredMixin,CreateView):
	model = Orden
	form_class = OrdenCreateForm
	template_name = 'orden/orden_form.html'

	def form_valid(self, form):
		self.object = form.save(self)
		if self.object:
			for component in self.object.componente.all():
				component.hDurg = 0
				component.mDurg = 0
				component.save()
			return super(ordenCreateView, self).form_valid(form)
		else:
			return self
	success_url = reverse_lazy('orden_list')

@method_decorator(staff_member_required, name='dispatch')
class ordenUpdateView(LoginRequiredMixin,UpdateView):
	model = Orden
	form_class = OrdenUpdateForm
	template_name = 'orden/orden_update_form.html'
	def get_success_url(self):
		return reverse_lazy('orden_detail', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class ordenDeleteView(LoginRequiredMixin,DeleteView):
	model = Orden
	template_name = 'orden/orden_confirm_delete.html'
	success_url = reverse_lazy('orden_list')

@method_decorator(staff_member_required, name='dispatch')
class ordenPDFListView(LoginRequiredMixin,PDFTemplateView):
	template_name = 'orden/pdf_orden_list.html'